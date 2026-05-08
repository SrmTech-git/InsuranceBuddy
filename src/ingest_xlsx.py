# ingest_xlsx.py — parse multi-state Excel spreadsheet into per-state document chunks

import openpyxl
from pathlib import Path
from langchain_core.documents import Document
from langchain_text_splitters import RecursiveCharacterTextSplitter
from states import STATE_MAP
from config import CHUNK_SIZE, CHUNK_OVERLAP

# Spreadsheet column headers use Title Case ("Ohio"); STATE_MAP keys are
# lowercase ("ohio") to match folder names. Derive once so adding a state
# in states.py is the only change needed.
_HEADER_TO_STATE: dict[str, str] = {name.title(): code for name, code in STATE_MAP.items()}

# Row types to skip when parsing data rows
_SKIP_DATA_POINTS = {"Data Point", None}


def _find_state_columns(ws, header_row: int = 4) -> dict[int, str]:
    """Return {col_index: state_code} for all state columns in the header row."""
    state_cols = {}
    for col in range(1, ws.max_column + 1):
        header = ws.cell(row=header_row, column=col).value
        if header in _HEADER_TO_STATE:
            state_cols[col] = _HEADER_TO_STATE[header]
    return state_cols


def _parse_quick_reference(ws, state_lines: dict[str, list[str]]) -> None:
    """Append the 'Quick Reference' sheet rows to each state's lines.

    Layout: row 1 = title, row 2 = header (col 0 = label, cols 1+ = state names),
    rows 3+ = one summary value per state per row. Mutates state_lines in place.
    """
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 3:
        return  # nothing usable

    header = rows[1]
    state_cols: dict[int, str] = {}
    for i, val in enumerate(header):
        if val and str(val).strip() in _HEADER_TO_STATE:
            state_cols[i] = _HEADER_TO_STATE[str(val).strip()]
    if not state_cols:
        return

    # Section break so the LLM sees a clear boundary
    for code in set(state_cols.values()):
        if code in state_lines:
            state_lines[code].append("\nQUICK REFERENCE")

    for row in rows[2:]:
        label = row[0]
        if not label:
            continue
        label_str = str(label).strip()
        for col_idx, code in state_cols.items():
            if col_idx >= len(row) or code not in state_lines:
                continue
            cell = row[col_idx]
            value = str(cell).strip() if cell is not None else "N/A"
            state_lines[code].append(f"  {label_str}: {value}")


def _find_notes_column(ws, header_row: int = 4) -> int | None:
    """Return the column index of 'Notes / Key Differences', or None."""
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row=header_row, column=col).value
        if val and "notes" in str(val).lower():
            return col
    return None


def load_xlsx_by_state(file_path: str) -> list[tuple[str, list[Document]]]:
    """Parse a multi-state reference spreadsheet into per-state document chunks.

    Returns a list of (state_code, chunks) pairs — one entry per state found
    in the spreadsheet. Each state's chunks carry:
        filename  : "<basename>_<STATE>"  (ensures dedup works per state)
        state     : two-letter state code
        description: human-readable label

    Only the 'State Auto Requirements' sheet is parsed.
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    if "State Auto Requirements" not in wb.sheetnames:
        raise ValueError(f"Expected sheet 'State Auto Requirements' in {file_path}")

    ws = wb["State Auto Requirements"]
    basename = Path(file_path).stem  # e.g. "StateSpreadSheet"

    # Locate header row — find the first row where col A = "Category"
    header_row = None
    for row in range(1, 10):
        if ws.cell(row=row, column=1).value == "Category":
            header_row = row
            break
    if header_row is None:
        raise ValueError("Could not find header row (Category column) in spreadsheet")

    state_cols = _find_state_columns(ws, header_row)
    notes_col = _find_notes_column(ws, header_row)

    if not state_cols:
        raise ValueError("No state columns found in spreadsheet header row")

    # Accumulate text lines per state
    state_lines: dict[str, list[str]] = {code: [] for code in state_cols.values()}
    current_section = ""
    current_category = ""

    for row in range(header_row + 1, ws.max_row + 1):
        cat_val = ws.cell(row=row, column=1).value
        dp_val = ws.cell(row=row, column=2).value

        # Section header row: col A has section title, col B is empty
        if cat_val and dp_val is None:
            # Could be a section header (e.g. "UM / UIM COVERAGE") or
            # a repeated "Category" header row — skip both
            if str(cat_val).strip().upper() != "CATEGORY":
                current_section = str(cat_val).strip()
                current_category = ""
                for code in state_lines:
                    state_lines[code].append(f"\n{current_section}")
            continue

        # Skip repeated column header rows
        if dp_val in _SKIP_DATA_POINTS:
            continue

        # Forward-fill category
        if cat_val:
            current_category = str(cat_val).strip()

        data_point = str(dp_val).strip() if dp_val else ""
        if not data_point:
            continue

        notes = ""
        if notes_col:
            note_val = ws.cell(row=row, column=notes_col).value
            if note_val:
                notes = f"  [Note: {str(note_val).strip()}]"

        # Write one line per state for this data point
        for col_idx, state_code in state_cols.items():
            cell_val = ws.cell(row=row, column=col_idx).value
            value = str(cell_val).strip() if cell_val is not None else "N/A"
            label = f"{current_category} — {data_point}" if current_category else data_point
            state_lines[state_code].append(f"  {label}: {value}{notes}")

    # Append Sheet 2 ("Quick Reference") if present — compact state-by-state
    # summary that complements the detailed Sheet 1 data. Sheet 3 ("Notes &
    # Sources") is intentionally skipped: it's maintainer disclaimers, not
    # retrieval material.
    if "Quick Reference" in wb.sheetnames:
        _parse_quick_reference(wb["Quick Reference"], state_lines)

    # Build one Document per state, then chunk it
    splitter = RecursiveCharacterTextSplitter(chunk_size=CHUNK_SIZE, chunk_overlap=CHUNK_OVERLAP)
    results: list[tuple[str, list[Document]]] = []

    for state_code, lines in state_lines.items():
        if not any(l.strip() for l in lines):
            continue

        state_name = next(
            (name for name, code in _HEADER_TO_STATE.items() if code == state_code),
            state_code,
        )
        header = (
            f"{state_name} ({state_code}) Auto Insurance Requirements\n"
            f"Source: {basename}.xlsx\n"
        )
        full_text = header + "\n".join(lines)

        base_doc = Document(
            page_content=full_text,
            metadata={
                "form_number": "",
                "edition_date": "",
                "description": f"{state_name} Auto Insurance Requirements",
                "filename": f"{basename}_{state_code}",
                "parsed": True,
                "state": state_code,
            },
        )
        chunks = splitter.split_documents([base_doc])
        results.append((state_code, chunks))

    return results


if __name__ == "__main__":
    import sys
    path = sys.argv[1] if len(sys.argv) > 1 else \
        r"C:\Users\shann\insurance-rag\data\raw\regulatory\reference\StateSpreadSheet.xlsx"

    state_docs = load_xlsx_by_state(path)
    for state_code, chunks in state_docs:
        print(f"\n{state_code}: {len(chunks)} chunk(s)")
        print(chunks[0].page_content[:300])
        print("...")
